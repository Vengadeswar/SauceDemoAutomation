import openpyxl


def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    lst = []

    for i in range(2, rows + 1):
        for j in range(1, cols + 1):
            data = sheet.cell(i, j).value
            lst.append(data)

    return lst
